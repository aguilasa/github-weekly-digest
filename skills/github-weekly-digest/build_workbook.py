#!/usr/bin/env python3
"""Turn a JSON dump of GitHub pull request / issue data into an Excel
workbook, and print an email-ready text summary.

This script does no networking of its own. The GitHub MCP connector
(mcp__github__list_pull_requests, mcp__github__search_pull_requests,
mcp__github__list_issues, mcp__github__search_issues) is used by Claude
to fetch the data first; this script only handles the data-to-workbook
transformation, so it can be tested with a fixture file independently
of any live MCP session (see references/sample-digest-data.json).

Expected input JSON shape:

{
  "repo": "owner/name",
  "days": 7,
  "pull_requests": [
    {"number": 1, "title": "...", "author": "...", "state": "open",
     "created_at": "...", "updated_at": "...", "merged_at": null,
     "url": "..."}
  ],
  "issues": [
    {"number": 2, "title": "...", "author": "...", "state": "closed",
     "created_at": "...", "updated_at": "...", "closed_at": "...",
     "url": "..."}
  ]
}
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class DigestError(Exception):
    """Raised for expected failures we want to report cleanly, not crash on."""


def log(message: str) -> None:
    print(f"[build-workbook] {message}", file=sys.stderr)


def load_data(path: Path) -> dict:
    if not path.exists():
        raise DigestError(f"input file not found: {path}")
    try:
        return json.loads(path.read_text())
    except json.JSONDecodeError as exc:
        raise DigestError(f"input file is not valid JSON: {path}") from exc


def build_workbook(data: dict) -> Workbook:
    workbook = Workbook()

    pr_sheet = workbook.active
    pr_sheet.title = "Pull Requests"
    _write_sheet(
        pr_sheet,
        headers=["#", "Title", "Author", "State", "Created", "Updated", "Merged", "URL"],
        rows=[
            [
                pr["number"],
                pr["title"],
                pr.get("author", "unknown"),
                pr["state"],
                pr.get("created_at", ""),
                pr.get("updated_at", ""),
                pr.get("merged_at") or "",
                pr.get("url", ""),
            ]
            for pr in data.get("pull_requests", [])
        ],
    )

    issue_sheet = workbook.create_sheet("Issues")
    _write_sheet(
        issue_sheet,
        headers=["#", "Title", "Author", "State", "Created", "Updated", "Closed", "URL"],
        rows=[
            [
                issue["number"],
                issue["title"],
                issue.get("author", "unknown"),
                issue["state"],
                issue.get("created_at", ""),
                issue.get("updated_at", ""),
                issue.get("closed_at") or "",
                issue.get("url", ""),
            ]
            for issue in data.get("issues", [])
        ],
    )

    workbook.properties.title = f"{data.get('repo', 'unknown repo')} weekly digest"
    return workbook


def _write_sheet(sheet, headers: list[str], rows: list[list]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    if not rows:
        sheet.append(["No activity in this period"])
    else:
        for row in rows:
            sheet.append(row)

    for idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(idx)
        max_len = max([len(header)] + [len(str(row[idx - 1])) for row in rows] or [0])
        sheet.column_dimensions[column_letter].width = min(max_len + 2, 60)


def summarize(data: dict) -> str:
    repo = data.get("repo", "unknown repo")
    days = data.get("days", 7)
    prs = data.get("pull_requests", [])
    issues = data.get("issues", [])

    open_prs = [pr for pr in prs if pr.get("state", "").lower() == "open"]
    merged_prs = [pr for pr in prs if pr.get("merged_at")]
    open_issues = [issue for issue in issues if issue.get("state", "").lower() == "open"]
    closed_issues = [issue for issue in issues if issue.get("state", "").lower() == "closed"]

    lines = [
        f"Weekly GitHub digest for {repo} (last {days} days)",
        "",
        f"Pull requests: {len(open_prs)} open, {len(merged_prs)} merged this period",
        f"Issues: {len(open_issues)} open, {len(closed_issues)} closed this period",
        "",
        "Full details attached in the Excel workbook.",
    ]
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("data", type=Path, help="path to the digest-data.json file")
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="output .xlsx path (default: output/digest-<repo>-<date>.xlsx)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    try:
        data = load_data(args.data)

        for required in ("repo",):
            if required not in data:
                raise DigestError(f"input JSON is missing required field: {required}")

        log(
            f"building workbook ({len(data.get('pull_requests', []))} PR rows, "
            f"{len(data.get('issues', []))} issue rows)..."
        )
        workbook = build_workbook(data)

        if args.out is not None:
            output_path = args.out
        else:
            output_dir = Path("output")
            output_dir.mkdir(exist_ok=True)
            today = datetime.now().strftime("%Y-%m-%d")
            safe_repo_name = data["repo"].replace("/", "-")
            output_path = output_dir / f"digest-{safe_repo_name}-{today}.xlsx"

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        log(f"workbook saved to {output_path}")

        print(summarize(data))
        print(f"WORKBOOK_PATH={output_path}")
        return 0

    except DigestError as exc:
        log(f"error: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
