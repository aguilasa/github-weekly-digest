#!/usr/bin/env python3
"""Fetch a GitHub repo's weekly PR/issue activity, write it to an Excel
workbook, and email a summary with the workbook attached.

Data retrieval uses the `gh` CLI (already authenticated on this machine),
so no separate GitHub token needs to be managed. Email is sent over
Gmail SMTP using an App Password.
"""

from __future__ import annotations

import argparse
import json
import os
import smtplib
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

PR_FIELDS = "number,title,author,state,createdAt,updatedAt,mergedAt,url"
ISSUE_FIELDS = "number,title,author,state,createdAt,updatedAt,closedAt,url"


class DigestError(Exception):
    """Raised for expected failures we want to report cleanly, not crash on."""


def log(message: str) -> None:
    print(f"[github-weekly-digest] {message}", file=sys.stderr)


def load_env_file(path: Path) -> None:
    """Load KEY=VALUE pairs from a .env file into os.environ, if present."""
    if not path.exists():
        return
    for raw_line in path.read_text().splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        os.environ.setdefault(key.strip(), value.strip())


def check_gh_available() -> None:
    try:
        subprocess.run(
            ["gh", "auth", "status"],
            check=True,
            capture_output=True,
            text=True,
        )
    except FileNotFoundError as exc:
        raise DigestError(
            "the `gh` CLI is not installed. Install it from https://cli.github.com/"
        ) from exc
    except subprocess.CalledProcessError as exc:
        raise DigestError(
            f"`gh` is not authenticated. Run `gh auth login` first.\n{exc.stderr.strip()}"
        ) from exc


def gh_json(args: list[str]) -> list[dict]:
    try:
        result = subprocess.run(
            ["gh", *args],
            check=True,
            capture_output=True,
            text=True,
        )
    except subprocess.CalledProcessError as exc:
        raise DigestError(f"`gh {' '.join(args)}` failed:\n{exc.stderr.strip()}") from exc

    try:
        return json.loads(result.stdout or "[]")
    except json.JSONDecodeError as exc:
        raise DigestError(f"could not parse `gh {' '.join(args)}` output as JSON") from exc


def fetch_pull_requests(repo: str, since: datetime) -> list[dict]:
    open_prs = gh_json(
        ["pr", "list", "--repo", repo, "--state", "open", "--json", PR_FIELDS, "--limit", "200"]
    )
    closed_prs = gh_json(
        ["pr", "list", "--repo", repo, "--state", "closed", "--json", PR_FIELDS, "--limit", "200"]
    )
    recent_closed = [
        pr for pr in closed_prs if _parse_date(pr.get("updatedAt")) >= since
    ]
    return open_prs + recent_closed


def fetch_issues(repo: str, since: datetime) -> list[dict]:
    open_issues = gh_json(
        ["issue", "list", "--repo", repo, "--state", "open", "--json", ISSUE_FIELDS, "--limit", "200"]
    )
    closed_issues = gh_json(
        ["issue", "list", "--repo", repo, "--state", "closed", "--json", ISSUE_FIELDS, "--limit", "200"]
    )
    recent_closed = [
        issue for issue in closed_issues if _parse_date(issue.get("updatedAt")) >= since
    ]
    return open_issues + recent_closed


def _parse_date(value: str | None) -> datetime:
    if not value:
        return datetime.min.replace(tzinfo=timezone.utc)
    return datetime.fromisoformat(value.replace("Z", "+00:00"))


def build_workbook(repo: str, prs: list[dict], issues: list[dict]) -> Workbook:
    workbook = Workbook()

    pr_sheet = workbook.active
    pr_sheet.title = "Pull Requests"
    _write_sheet(
        pr_sheet,
        headers=["#", "Title", "Author", "State", "Created", "Updated", "Merged", "URL"],
        rows=[
            [
                pr["number"],
                pr["title"],
                pr.get("author", {}).get("login", "unknown"),
                pr["state"],
                pr.get("createdAt", ""),
                pr.get("updatedAt", ""),
                pr.get("mergedAt") or "",
                pr.get("url", ""),
            ]
            for pr in prs
        ],
    )

    issue_sheet = workbook.create_sheet("Issues")
    _write_sheet(
        issue_sheet,
        headers=["#", "Title", "Author", "State", "Created", "Updated", "Closed", "URL"],
        rows=[
            [
                issue["number"],
                issue["title"],
                issue.get("author", {}).get("login", "unknown"),
                issue["state"],
                issue.get("createdAt", ""),
                issue.get("updatedAt", ""),
                issue.get("closedAt") or "",
                issue.get("url", ""),
            ]
            for issue in issues
        ],
    )

    workbook.properties.title = f"{repo} weekly digest"
    return workbook


def _write_sheet(sheet, headers: list[str], rows: list[list]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    if not rows:
        sheet.append(["No activity in this period"])
    else:
        for row in rows:
            sheet.append(row)

    for idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(idx)
        max_len = max([len(header)] + [len(str(row[idx - 1])) for row in rows] or [0])
        sheet.column_dimensions[column_letter].width = min(max_len + 2, 60)


def summarize(repo: str, days: int, prs: list[dict], issues: list[dict]) -> str:
    open_prs = [pr for pr in prs if pr["state"] == "OPEN"]
    merged_prs = [pr for pr in prs if pr.get("mergedAt")]
    open_issues = [issue for issue in issues if issue["state"] == "OPEN"]
    closed_issues = [issue for issue in issues if issue["state"] == "CLOSED"]

    lines = [
        f"Weekly GitHub digest for {repo} (last {days} days)",
        "",
        f"Pull requests: {len(open_prs)} open, {len(merged_prs)} merged this period",
        f"Issues: {len(open_issues)} open, {len(closed_issues)} closed this period",
        "",
        "Full details attached in the Excel workbook.",
    ]
    return "\n".join(lines)


def send_email(subject: str, body: str, attachment_path: Path) -> None:
    email_from = os.environ.get("EMAIL_FROM")
    app_password = os.environ.get("EMAIL_APP_PASSWORD")
    email_to = os.environ.get("EMAIL_TO")

    missing = [
        name
        for name, value in (
            ("EMAIL_FROM", email_from),
            ("EMAIL_APP_PASSWORD", app_password),
            ("EMAIL_TO", email_to),
        )
        if not value
    ]
    if missing:
        raise DigestError(
            f"missing required email settings in .env: {', '.join(missing)}"
        )

    message = MIMEMultipart()
    message["From"] = email_from
    message["To"] = email_to
    message["Subject"] = subject
    message.attach(MIMEText(body, "plain"))

    with attachment_path.open("rb") as handle:
        attachment = MIMEApplication(handle.read(), Name=attachment_path.name)
    attachment["Content-Disposition"] = f'attachment; filename="{attachment_path.name}"'
    message.attach(attachment)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(email_from, app_password)
            server.send_message(message)
    except smtplib.SMTPException as exc:
        raise DigestError(f"failed to send email: {exc}") from exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--repo",
        default=os.environ.get("REPO"),
        help="owner/repo to digest, e.g. aguilasa/css-windify",
    )
    parser.add_argument(
        "--days",
        type=int,
        default=int(os.environ.get("DAYS", "7")),
        help="how many days back to include (default: 7)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="build the workbook but skip sending the email",
    )
    return parser.parse_args()


def main() -> int:
    project_root = Path(__file__).resolve().parent.parent.parent
    load_env_file(project_root / ".env")

    args = parse_args()

    if not args.repo:
        log("error: no --repo given and REPO is not set in .env")
        return 2

    try:
        check_gh_available()

        since = datetime.now(timezone.utc) - timedelta(days=args.days)

        log(f"fetching pull requests for {args.repo}...")
        prs = fetch_pull_requests(args.repo, since)

        log(f"fetching issues for {args.repo}...")
        issues = fetch_issues(args.repo, since)

        log(f"building workbook ({len(prs)} PR rows, {len(issues)} issue rows)...")
        workbook = build_workbook(args.repo, prs, issues)

        output_dir = project_root / "output"
        output_dir.mkdir(exist_ok=True)
        today = datetime.now().strftime("%Y-%m-%d")
        safe_repo_name = args.repo.replace("/", "-")
        output_path = output_dir / f"digest-{safe_repo_name}-{today}.xlsx"
        workbook.save(output_path)
        log(f"workbook saved to {output_path}")

        summary = summarize(args.repo, args.days, prs, issues)
        print(summary)

        if args.dry_run:
            log("dry run: skipping email send")
            return 0

        log("sending email...")
        send_email(
            subject=f"GitHub weekly digest: {args.repo}",
            body=summary,
            attachment_path=output_path,
        )
        log(f"email sent to {os.environ.get('EMAIL_TO')}")
        return 0

    except DigestError as exc:
        log(f"error: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
